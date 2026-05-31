import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


df = pd.read_csv(r'E:\superstore\superstore_cleaned.csv', encoding='latin-1')

wb = openpyxl.Workbook()

ws = wb.active
ws.title = 'Sales Report'


headers = ['Order ID', 'Customer Name', 'Segment', 'Region',
           'City', 'State', 'Category', 'Sub-Category',
           'Sales', 'Ship Mode', 'Order Date', 'Shipping Days',
           'Sales Category', 'Order Year', 'Order Month']

for col,header in enumerate(headers,1):
    ws.cell(row=1,column=col,value=header)


for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = PatternFill(start_color='1A1A2E', 
                            end_color='1A1A2E', 
                            fill_type='solid')
    cell.font = Font(color='FFFFFF', bold=True)
    cell.alignment = Alignment(horizontal='center')

for i,x in enumerate(df[headers].values,2):
    for colu,value in enumerate(x,1):
        ws.cell(row=i,column=colu,value=value)


for row in range(2, len(df) + 2):
    cell = ws.cell(row=row, column=13)
    if cell.value == 'High':
        cell.fill = PatternFill(start_color='90EE90',
                               end_color='90EE90',
                               fill_type='solid')
    elif cell.value == 'Medium':
        cell.fill = PatternFill(start_color='FFB347',
                               end_color='FFB347',
                               fill_type='solid')
    elif cell.value == 'Low':
        cell.fill = PatternFill(start_color='FF6B6B',
                               end_color='FF6B6B',
                               fill_type='solid')

ws_summary = wb.create_sheet('Summary')

region_summary = df.groupby('Region')['Sales'].sum().round(2).reset_index()

sales_category = df.groupby('Category')['Sales'].sum().round().reset_index()

order_sales_year = df.groupby('Order Year')['Sales'].sum().round().reset_index()


# Region Summary
ws_summary.cell(row=1, column=1, value='Sales by Region')
ws_summary.cell(row=2, column=1, value='Region')
ws_summary.cell(row=2, column=2, value='Total Sales')

for idx, row in enumerate(region_summary.values, 3):
    ws_summary.cell(row=idx, column=1, value=row[0])
    ws_summary.cell(row=idx, column=2, value=row[1])

# Category Summary
ws_summary.cell(row=10, column=1, value='Sales by Category')
ws_summary.cell(row=11, column=1, value='Category')
ws_summary.cell(row=11, column=2, value='Total Sales')

for idx, row in enumerate(sales_category.values, 12):
    ws_summary.cell(row=idx, column=1, value=row[0])
    ws_summary.cell(row=idx, column=2, value=row[1])

# Year Summary
ws_summary.cell(row=18, column=1, value='Sales by Year')
ws_summary.cell(row=19, column=1, value='Year')
ws_summary.cell(row=19, column=2, value='Total Sales')

for idx, row in enumerate(order_sales_year.values, 20):
    ws_summary.cell(row=idx, column=1, value=row[0])
    ws_summary.cell(row=idx, column=2, value=row[1])


wb.save(r'E:\superstore\superstore_report.xlsx')
print("✅ Report created successfully!")
